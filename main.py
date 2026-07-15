import openpyxl
from openpyxl.styles import PatternFill
import pyautogui
import pyperclip
import time
import os

# Trava de segurança: arraste o mouse para o canto da tela para abortar
pyautogui.FAILSAFE = True

AMARELO = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
AZUL = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

def clicar_na_imagem(nome_imagem, cliques=1, esperar=1.5, confidence=0.85):
    """Procura a imagem com 85% de precisão e clica"""
    try:
        caminho_imagem = os.path.join(os.getcwd(), nome_imagem)
        if not os.path.exists(caminho_imagem):
            print(f"[!] Aviso: Imagem '{nome_imagem}' não encontrada!")
            return False
        
        ponto = pyautogui.locateCenterOnScreen(caminho_imagem, confidence=confidence)
        if ponto:
            pyautogui.moveTo(ponto.x, ponto.y, duration=0.4)
            time.sleep(0.3) # PAUSA ESTRITA: Dá tempo para o site registrar que o mouse está em cima
            if cliques == 1:
                pyautogui.click()
            elif cliques == 2:
                pyautogui.doubleClick()
            time.sleep(esperar)
            return True
        return False
    except Exception as e:
        return False

def procurar_imagem_na_tela(nome_imagem, confidence=0.85):
    """Apenas verifica se a imagem existe na tela"""
    caminho = os.path.join(os.getcwd(), nome_imagem)
    if os.path.exists(caminho):
        return pyautogui.locateOnScreen(caminho, confidence=confidence) is not None
    return False

def rolar_e_procurar(nome_imagem, direcao="baixo", tentativas=6, confidence=0.85):
    """Rola a tela aos pouquinhos (scroll inteligente) até achar o botão"""
    for _ in range(tentativas):
        if procurar_imagem_na_tela(nome_imagem, confidence=confidence):
            return True # Achou!
        if direcao == "baixo":
            pyautogui.scroll(-300) # Rola um pouco para baixo
        else:
            pyautogui.scroll(300)  # Rola um pouco para cima
        time.sleep(1)
    return False # Desiste se rolar 6 vezes e não achar

def processar_planilha():
    arquivo_excel = 'Ocorrencias_MID.xlsx'
    
    if not os.path.exists(arquivo_excel):
        print(f"Erro crítico: O arquivo '{arquivo_excel}' não foi encontrado!")
        input("Pressione Enter para sair...")
        return

    wb = openpyxl.load_workbook(arquivo_excel)
    planilha = wb.active

    print("\n========================================================")
    print("ROBÔ INICIANDO EM 5 SEGUNDOS...")
    print("1. Feche a planilha (se estiver aberta).")
    print("2. Puxe o Chrome para o MONITOR PRINCIPAL.")
    print("3. Clique na tela do Chrome para focar.")
    print("========================================================\n")
    time.sleep(5)

    for linha in range(2, planilha.max_row + 1):
        apolice = planilha.cell(row=linha, column=1).value
        endosso = planilha.cell(row=linha, column=2).value
        ocorrencia = planilha.cell(row=linha, column=5).value
        
        cor_atual = planilha.cell(row=linha, column=1).fill.start_color.index

        if not apolice:
            break

        if type(cor_atual) == str and ('FFFF00' in cor_atual or '00B0F0' in cor_atual):
            continue

        if endosso is not None and str(endosso).strip() != "":
            continue

        ocorrencias_alvo = [
            "SOMA DOS PRÊMIOS DAS PARCS NÃO BATE COM O PRÊMIO TOTAL",
            "INÍCIO DE VIGÊNCIA ANTERIOR A 30 DIAS DA DATA DA PROPOSTA",
            "INÍCIO DE VIGÊNCIA ANTERIOR A 30 DIAS DA DATA DE EMISSÃO"
        ]

        if ocorrencia in ocorrencias_alvo:
            print(f"\n[{linha}] Iniciando Apólice: {apolice}")
            pyperclip.copy(str(apolice))
            
            # --- PESQUISA A APÓLICE ---
            if not clicar_na_imagem('campo_pesquisa.png', cliques=1):
                planilha.cell(row=linha, column=1).fill = AZUL
                continue
            
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.2)
            pyautogui.press('backspace')
            time.sleep(0.2)
            pyautogui.hotkey('ctrl', 'v')
            pyautogui.press('enter')
            time.sleep(4) 
            
            clicar_na_imagem('icone_abrir_apolice.png', esperar=3)

            # Tira o mouse do caminho
            pyautogui.move(300, 0, duration=0.3)
            time.sleep(1)

            # --- ETAPA DE CHECK ESTRITO 1: ACHAR OCORRÊNCIAS ---
            print(f"[{linha}] Procurando o botão Ocorrências...")
            if rolar_e_procurar('botao_ocorrencias.png', direcao="baixo"):
                clicar_na_imagem('botao_ocorrencias.png', esperar=2)
            else:
                print(f"[{linha}] ERRO: Não encontrou o botão de ocorrências. Pulando.")
                planilha.cell(row=linha, column=1).fill = AZUL
                pyautogui.scroll(2000)
                clicar_na_imagem('logomarca.png', esperar=3)
                continue # Aborta esta apólice e vai para a próxima

            # --- ETAPA DE CHECK ESTRITO 2: VALIDAR SE ABRIU ---
            print(f"[{linha}] Validando se a aba abriu (procurando Status 1)...")
            pyautogui.scroll(-300) # Rola um chorinho pra garantir que o painel está visível
            time.sleep(1)
            
            if not procurar_imagem_na_tela('status_1.png', confidence=0.85):
                print(f"[{linha}] ERRO: A aba não abriu ou o Status 1 não foi encontrado. Abortando.")
                planilha.cell(row=linha, column=1).fill = AZUL
                pyautogui.scroll(2000)
                clicar_na_imagem('logomarca.png', esperar=3)
                continue # Pula se falhou no check

            # =======================================================
            # DAQUI PRA BAIXO, ELE SÓ EXECUTA SE PASSOU NOS CHECKS
            # =======================================================
            sucesso_cliques = False

            if ocorrencia == "SOMA DOS PRÊMIOS DAS PARCS NÃO BATE COM O PRÊMIO TOTAL":
                if rolar_e_procurar('aba_premios.png', direcao="baixo"):
                    clicar_na_imagem('aba_premios.png', esperar=2)
                    clicar_na_imagem('data_vencimento.png')
                    clicar_na_imagem('clicar_fora.png')
                    
                    if not procurar_imagem_na_tela('gravar.png'):
                        pyautogui.scroll(-500) 
                        time.sleep(1)
                    
                    if clicar_na_imagem('gravar.png', esperar=4):
                        sucesso_cliques = True

            elif ocorrencia in ["INÍCIO DE VIGÊNCIA ANTERIOR A 30 DIAS DA DATA DA PROPOSTA", 
                                "INÍCIO DE VIGÊNCIA ANTERIOR A 30 DIAS DA DATA DE EMISSÃO"]:
                pyautogui.scroll(2500) # Sobe com força pra garantir que está no topo
                time.sleep(1.5)
                
                if clicar_na_imagem('data_proposta.png'):
                    clicar_na_imagem('clicar_fora.png')
                    
                    if not procurar_imagem_na_tela('gravar.png'):
                        pyautogui.scroll(-500)
                        time.sleep(1)
                        
                    if clicar_na_imagem('gravar.png', esperar=4):
                        sucesso_cliques = True

            # --- VERIFICAÇÃO FINAL ---
            pyautogui.scroll(2000) 
            time.sleep(1)
            # Rola aos poucos até achar o ocorrências de novo para olhar se saiu do 1
            rolar_e_procurar('botao_ocorrencias.png', direcao="baixo")
            clicar_na_imagem('botao_ocorrencias.png', esperar=2)
            pyautogui.scroll(-300)
            time.sleep(1.5)
            
            ocorrencia_ainda_ativa = procurar_imagem_na_tela('status_1.png', confidence=0.85)

            if sucesso_cliques and not ocorrencia_ainda_ativa:
                planilha.cell(row=linha, column=1).fill = AMARELO
            else:
                planilha.cell(row=linha, column=1).fill = AZUL
                
            wb.save('Ocorrencias_MID_Atualizado.xlsx')
            
            # --- VOLTAR AO INÍCIO ---
            pyautogui.scroll(2500) 
            time.sleep(1.5)
            clicar_na_imagem('logomarca.png', esperar=3)

    print("\nPROCESSO FINALIZADO! Pode abrir o arquivo Ocorrencias_MID_Atualizado.xlsx.")
    input("Pressione Enter para fechar...")

if __name__ == "__main__":
    processar_planilha()
